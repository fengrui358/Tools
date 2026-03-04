"""
诊断工具：分析 Excel 文件的单元格边框情况
用于调试合并问题
"""
import sys
from openpyxl import load_workbook
from openpyxl.styles import Border

sys.stdout.reconfigure(encoding='utf-8')

def _has_bottom_border(cell) -> bool:
    """检查单元格是否有下边框"""
    if cell.border is None:
        return False
    border_attr = getattr(cell.border, 'bottom', None)
    if border_attr is None:
        return False
    return border_attr.style is not None and border_attr.style != 'none'

def diagnose(input_file: str, sheet_name: str = "Sheet1",
             start_row: int = 25, end_row: int = 45,
             col: int = 6):
    """
    诊断指定范围内单元格的边框情况

    Args:
        input_file: 输入文件路径
        sheet_name: 工作表名称
        start_row: 开始行
        end_row: 结束行
        col: 列号（1-based）
    """
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    col_letter = chr(64 + col) if col <= 26 else f"?{col}"

    print(f"\n诊断文件: {input_file}")
    print(f"工作表: {sheet_name}")
    print(f"列: {col} ({col_letter})")
    print(f"行范围: {start_row}-{end_row}")
    print("\n" + "=" * 80)

    for row_idx in range(start_row, end_row + 1):
        cell = ws.cell(row=row_idx, column=col)
        has_bottom = _has_bottom_border(cell)
        value = str(cell.value) if cell.value else "(空)"

        # 边框标记
        border_mark = "┄" if has_bottom else " "

        # 内容显示（限制长度）
        if len(value) > 40:
            value = value[:37] + "..."

        print(f"行 {row_idx:3d} [{border_mark}] {value}")

        # 如果有下边框，画一条横线
        if has_bottom:
            print("      " + "─" * 76)

    print("=" * 80)
    print("\n说明: ┄ = 有下边框, 空白 = 无下边框")
    print("Word 拆分特征: 中间行无下边框，最后一行有下边框\n")

    wb.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: uv run python diagnose.py <excel文件> [起始行] [结束行] [列号]")
        sys.exit(1)

    input_file = sys.argv[1]
    start_row = int(sys.argv[2]) if len(sys.argv) > 2 else 25
    end_row = int(sys.argv[3]) if len(sys.argv) > 3 else 45
    col = int(sys.argv[4]) if len(sys.argv) > 4 else 6

    diagnose(input_file, start_row=start_row, end_row=end_row, col=col)
