"""
Word2Excel Fixer 核心模块

实现 Excel 表格修复逻辑：通过边框检测合并被拆分的单元格
"""

import sys
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List, Set, Tuple
import logging

# 配置日志
if sys.platform == 'win32':
    try:
        # 在 Windows 上尝试使用 UTF-8 编码
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except (AttributeError, LookupError):
        pass

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


def _has_full_border(cell, border_type: str) -> bool:
    """
    检查单元格的指定边是否有完整边框

    Args:
        cell: openpyxl 单元格对象
        border_type: 边框类型 ('top', 'bottom', 'left', 'right')

    Returns:
        bool: 是否有完整边框
    """
    if cell.border is None:
        return False

    border_attr = getattr(cell.border, border_type, None)
    if border_attr is None:
        return False

    # 检查边框样式是否有效（不为 None 且不为 'none'）
    return border_attr.style is not None and border_attr.style != 'none'


def _has_top_border(cell) -> bool:
    """检查单元格是否有上边框"""
    return _has_full_border(cell, 'top')


def _has_bottom_border(cell) -> bool:
    """检查单元格是否有下边框"""
    return _has_full_border(cell, 'bottom')


def _is_empty_row(ws, row_idx: int, max_col: int) -> bool:
    """
    检查一行是否为空行（所有单元格都没有内容或只有空白字符）

    Args:
        ws: 工作表对象
        row_idx: 行索引（从 1 开始）
        max_col: 最大列数

    Returns:
        bool: 是否为空行
    """
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and str(cell.value).strip():
            return False
    return True


def _find_merge_groups_by_column(ws, max_row: int, max_col: int) -> dict:
    """
    找出每列需要合并的行组

    核心逻辑：遍历每一列，找出下边框缺失的连续行，这些行需要合并。

    从Word复制到Excel时，被拆分的单元格特征是：
    - 中间行没有下边框（单元格被拆分）
    - 最后一行有下边框（原始Word单元格的结束边界）

    **重要**：每列独立检测合并组，因为不同列的合并边界可能不同。

    Args:
        ws: 工作表对象
        max_row: 最大行数
        max_col: 最大列数

    Returns:
        dict: {col_idx: [[start_row, end_row], ...]} 每列的合并组列表
    """
    merge_groups_by_col = {}

    # 按列遍历，找出需要合并的行
    for col_idx in range(1, max_col + 1):
        merge_groups = []
        row_idx = 1

        while row_idx <= max_row:
            cell = ws.cell(row=row_idx, column=col_idx)

            # 找到有内容的行作为合并组开始
            # 注意：这里不跳过空行，因为空行也可能是合并组的一部分
            if not cell.value:
                row_idx += 1
                continue

            # 如果当前行没有下边框，可能是被拆分单元格的开始
            if not _has_bottom_border(cell):
                group_start = row_idx
                group_end = row_idx

                # 向下查找，直到找到有下边框的行
                for next_row in range(row_idx + 1, max_row + 1):
                    next_cell = ws.cell(row=next_row, column=col_idx)

                    # 如果遇到空单元格，继续查找（可能是合并组的中间行）
                    # 但如果连续多个空单元格，可能到了表格末尾
                    if not next_cell.value:
                        # 检查是否后面还有非空单元格，如果有则继续
                        has_more_content = False
                        for check_row in range(next_row + 1, min(next_row + 5, max_row + 1)):
                            if ws.cell(row=check_row, column=col_idx).value:
                                has_more_content = True
                                break
                        if not has_more_content:
                            break

                    # 找到有下边框的行，这是合并组的结束
                    if _has_bottom_border(next_cell):
                        group_end = next_row
                        break

                    # 没有下边框，继续作为同一组
                    group_end = next_row

                # 如果找到多行需要合并，添加到该列的合并组
                if group_end > group_start:
                    merge_groups.append([group_start, group_end])
                    row_idx = group_end + 1
                else:
                    row_idx += 1
            else:
                row_idx += 1

        if merge_groups:
            merge_groups_by_col[col_idx] = merge_groups

    return merge_groups_by_col


def _merge_overlapping_groups(groups: List[List[int]]) -> List[List[int]]:
    """
    合并重叠或相邻的行组

    Args:
        groups: 行组列表，每个组是 [start, end]

    Returns:
        合并后的行组列表
    """
    if not groups:
        return []

    # 按起始行排序
    sorted_groups = sorted(groups, key=lambda x: x[0])

    merged = [sorted_groups[0]]
    for current in sorted_groups[1:]:
        last = merged[-1]

        # 如果当前组与上一组重叠或相邻，合并它们
        if current[0] <= last[1] + 1:
            last[1] = max(last[1], current[1])
        else:
            merged.append(current)

    return merged


def _merge_rows_content(ws, group: List[int], max_col: int) -> str:
    """
    合并同一组中所有行的内容，用换行符连接

    Args:
        ws: 工作表对象
        group: 行组 [start_row, end_row]
        max_col: 最大列数

    Returns:
        合并后的内容字典 {col_idx: merged_content}
    """
    merged_content = {}

    for col_idx in range(1, max_col + 1):
        contents = []
        for row_idx in range(group[0], group[1] + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                contents.append(str(cell.value).strip())

        if contents:
            # 用换行符连接多行内容
            merged_content[col_idx] = '\n'.join(contents)

    return merged_content


def _identify_key_columns(ws, max_row: int, max_col: int) -> List[int]:
    """
    识别主键列（包含分组标识的列）

    主键列的特征：
    - 每个合并组的第一行有内容（如服务分类名称）
    - 后续行为空或重复内容
    - 该列的内容密度适中（不是太高也不是太低）

    Args:
        ws: 工作表对象
        max_row: 最大行数
        max_col: 最大列数

    Returns:
        主键列的列表（按优先级排序）
    """
    key_columns = []

    for col_idx in range(1, max_col + 1):
        # 统计非空单元格的数量和分组模式
        non_empty_count = 0
        group_starts = 0  # 分组开始的数量（有内容且后有空行的位置）
        consecutive_empty = 0  # 连续空行的数量

        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            has_content = cell.value and str(cell.value).strip()

            if has_content:
                non_empty_count += 1

                # 检查是否是分组开始（当前有内容，且下一行是空）
                if row_idx < max_row:
                    next_cell = ws.cell(row=row_idx + 1, column=col_idx)
                    if not next_cell.value or not str(next_cell.value).strip():
                        group_starts += 1
            else:
                consecutive_empty += 1

        # 计算密度（非空行占总行数的比例）
        density = non_empty_count / max_row if max_row > 0 else 0

        # 主键列的特征：
        # - 密度适中（5%-50%之间），太高是描述列，太低是稀疏列
        # - 有较多的分组开始位置
        # - 有较多连续空行（表明是分组结构）

        # 计算评分：分组开始数量权重最高，连续空行次之，密度适中再次
        if 0.05 <= density <= 0.5 and group_starts >= 1:
            # 综合评分
            score = (
                group_starts * 200 +  # 分组开始数量权重最高
                consecutive_empty * 10 +  # 连续空行数量
                (1 - abs(density - 0.2)) * 50  # 密度越接近20%越好
            )
            key_columns.append((col_idx, score, density, group_starts))

    # 按评分排序
    key_columns.sort(key=lambda x: x[1], reverse=True)

    # 输出调试信息
    if key_columns:
        debug_info = []
        for col, score, density, groups in key_columns[:5]:
            debug_info.append(f"列{col}(密度{density:.1%},分组{groups})")
        logger.debug(f"候选主键列: {debug_info}")

    result = [col[0] for col in key_columns[:3]]

    # 如果仍然没有找到，使用内容密度适中的列
    if not result:
        for col_idx in range(1, max_col + 1):
            non_empty = sum(1 for r in range(1, max_row + 1)
                          if ws.cell(row=r, column=col_idx).value)
            density = non_empty / max_row
            if 0.1 <= density <= 0.7:
                result.append(col_idx)
                if len(result) >= 3:
                    break

    # 如果还是没有，使用第一列
    if not result and max_col >= 1:
        result = [1]

    return result


def _find_merge_groups_by_key_column(ws, max_row: int, max_col: int, key_col: int) -> List[List[int]]:
    """
    基于主键列找出所有需要合并的行组

    Args:
        ws: 工作表对象
        max_row: 最大行数
        max_col: 最大列数
        key_col: 主键列索引

    Returns:
        合并组列表 [[start_row, end_row], ...]
    """
    merge_groups = []
    row_idx = 1

    while row_idx <= max_row:
        key_cell = ws.cell(row=row_idx, column=key_col)

        # 找到有内容的主键行，作为合并组的开始
        if not key_cell.value or not str(key_cell.value).strip():
            row_idx += 1
            continue

        group_start = row_idx
        group_end = row_idx

        # 向下查找这个合并组的结束
        # 结束条件：遇到下一个有主键内容的行，或者遇到表格末尾
        for next_row in range(row_idx + 1, max_row + 1):
            next_key_cell = ws.cell(row=next_row, column=key_col)

            # 如果下一行有主键内容，说明是新组的开始
            if next_key_cell.value and str(next_key_cell.value).strip():
                break

            # 检查其他列是否有下边框（标识合并组的结束）
            has_bottom_border = False
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=next_row, column=col_idx)
                if _has_bottom_border(cell):
                    has_bottom_border = True
                    break

            if has_bottom_border:
                group_end = next_row
                break

            group_end = next_row

        # 如果找到多行需要合并，添加到合并组
        if group_end > group_start:
            merge_groups.append([group_start, group_end])
            row_idx = group_end + 1
        else:
            row_idx += 1

    return merge_groups


def _process_worksheet(ws, anchor_column: Optional[int] = None) -> Tuple[int, int]:
    """
    处理单个工作表，执行合并操作

    基于锚定列检测和合并被拆分的单元格。

    Args:
        ws: 工作表对象
        anchor_column: 锚定列索引（1-based），如果为None则自动识别

    Returns:
        (merged_groups, deleted_rows): 合并的组数和删除的行数
    """
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2 or max_col < 1:
        logger.info(f"工作表 '{ws.title}' 数据量不足，跳过处理")
        return 0, 0

    logger.info(f"处理工作表 '{ws.title}': {max_row} 行 x {max_col} 列")

    # 确定锚定列
    if anchor_column is not None:
        key_col = anchor_column
        logger.info(f"使用指定的锚定列: 列{key_col} ({get_column_letter(key_col)})")
    else:
        # 自动识别主键列
        key_columns = _identify_key_columns(ws, max_row, max_col)
        logger.debug(f"候选主键列: {key_columns}")

        if not key_columns:
            logger.info(f"工作表 '{ws.title}' 未找到主键列，跳过处理")
            return 0, 0

        key_col = key_columns[0]

    # 基于锚定列找出合并组
    merge_groups = _find_merge_groups_by_key_column(ws, max_row, max_col, key_col)

    if not merge_groups:
        logger.info(f"工作表 '{ws.title}' 未发现需要合并的行")
        return 0, 0

    logger.info(f"发现 {len(merge_groups)} 组需要合并的行（基于列 {key_col}）")

    # 记录需要删除的行
    rows_to_delete: Set[int] = set()

    # 处理每个合并组
    for group in merge_groups:
        start_row, end_row = group

        if end_row <= start_row:
            continue

        logger.debug(f"合并行组: 第 {start_row}-{end_row} 行")

        # 合并所有列的内容
        for col_idx in range(1, max_col + 1):
            contents = []
            for row_idx in range(start_row, end_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 跳过已合并的单元格（只读）
                if isinstance(cell, MergedCell):
                    continue
                if cell.value:
                    contents.append(str(cell.value).strip())

            if contents:
                # 合并内容到第一行
                target_cell = ws.cell(row=start_row, column=col_idx)
                if not isinstance(target_cell, MergedCell):
                    target_cell.value = '\n'.join(contents)

                # 清空被合并的行（保留第一行）
                for row_idx in range(start_row + 1, end_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not isinstance(cell, MergedCell):
                        cell.value = None

        # 标记需要删除的行（稍后统一处理）
        for row_idx in range(start_row + 1, end_row + 1):
            # 只有当整行都是空的时候才删除
            if _is_empty_row(ws, row_idx, max_col):
                rows_to_delete.add(row_idx)

    # 删除多余的行（倒序删除，避免行号变化）
    deleted_rows = 0
    if rows_to_delete:
        # 排序后倒序删除
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)
            deleted_rows += 1

    return len(merge_groups), deleted_rows


def fix_excel_from_file(input_path: str, output_path: Optional[str] = None,
                       anchor_column: int = 1) -> str:
    """
    修复从 Word 复制到 Excel 的表格文件

    Args:
        input_path: 输入的 Excel 文件路径
        output_path: 输出文件路径，默认为输入文件同目录下的 *_fixed.xlsx
        anchor_column: 锚定列索引（1-based），默认为1（第一列）

    Returns:
        str: 输出文件的完整路径

    Raises:
        FileNotFoundError: 输入文件不存在
        Exception: 文件处理过程中的其他错误
    """
    input_file = Path(input_path)

    # 检查输入文件是否存在
    if not input_file.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    # 检查文件扩展名
    if input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"不支持的文件格式: {input_file.suffix}，仅支持 .xlsx 和 .xlsm")

    # 生成输出文件路径
    if output_path is None:
        output_path = input_file.parent / f"{input_file.stem}_fixed{input_file.suffix}"
    else:
        output_path = Path(output_path)

    logger.info(f"开始处理文件: {input_path}")
    logger.info(f"输出文件: {output_path}")
    logger.info(f"使用锚定列: 列{anchor_column} ({get_column_letter(anchor_column)})")

    try:
        # 加载工作簿
        wb = load_workbook(input_file)

        total_groups = 0
        total_deleted = 0
        processed_sheets = 0

        # 处理所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            groups, deleted = _process_worksheet(ws, anchor_column)
            total_groups += groups
            total_deleted += deleted
            if groups > 0:
                processed_sheets += 1

        # 保存结果
        wb.save(output_path)
        wb.close()

        logger.info(f"处理完成！")
        logger.info(f"  处理工作表数: {processed_sheets}")
        logger.info(f"  合并行组数: {total_groups}")
        logger.info(f"  删除行数: {total_deleted}")
        logger.info(f"  输出文件: {output_path}")

        return str(output_path)

    except Exception as e:
        logger.error(f"处理文件时出错: {e}")
        raise


def fix_excel(input_path: str, output_path: Optional[str] = None) -> str:
    """
    fix_excel_from_file 的别名，提供更简洁的调用方式

    Args:
        input_path: 输入的 Excel 文件路径
        output_path: 输出文件路径，默认为输入文件同目录下的 *_fixed.xlsx

    Returns:
        str: 输出文件的完整路径
    """
    return fix_excel_from_file(input_path, output_path)
